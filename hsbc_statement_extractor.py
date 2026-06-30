from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
import pdfplumber
import pytesseract
from openpyxl.styles import Alignment, Font, PatternFill

HSBC_COLUMN_ORDER = [
    "Posting date",
    "transection Date",
    "Transection details",
    "Amount (AED)",
]

SHORT_DATE_RE = re.compile(r"^\d{2}[A-Z]{3}$")
AMOUNT_RE = re.compile(r"(-?[\d.,]+)\s*(CR)?$", re.IGNORECASE)
SUMMARY_MARKERS = ("ACCOUNT SUMMARY", "PAYMENT SUMMARY", "ISSUED BY HSBC")
MONTHS = ("JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC")

OCR_DIGIT_MAP = str.maketrans({
    "O": "0",
    "Q": "0",
    "D": "0",
    "I": "1",
    "L": "1",
    "|": "1",
    "]": "1",
    "[": "1",
    "S": "5",
    "B": "8",
})

OCR_MONTH_CHAR_MAP = str.maketrans({
    "0": "O",
    "1": "J",
    "I": "J",
    "L": "J",
    "]": "J",
    "[": "J",
    "7": "J",
    "4": "A",
    "5": "S",
    "8": "B",
    "T": "J",
})


def normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_ocr_text(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    for old, new in {
        "_": " ",
        "‘": "",
        "’": "",
        "“": "",
        "”": "",
        "—": " ",
        "–": " ",
        "|": " ",
        ">": " ",
    }.items():
        ascii_text = ascii_text.replace(old, new)
    return normalize_space(ascii_text.upper())


def normalize_date_token(token: str) -> str:
    compact = re.sub(r"[^A-Z0-9\]\[]+", "", normalize_ocr_text(token))
    if len(compact) == 6 and compact[0] in {"0", "O"}:
        compact = compact[1:]
    if len(compact) != 5:
        return compact

    day = compact[:2].translate(OCR_DIGIT_MAP)
    month_raw = compact[2:].translate(OCR_MONTH_CHAR_MAP)
    if not day.isdigit():
        return compact

    best_month = month_raw
    best_score = -1
    for month in MONTHS:
        score = sum(left == right for left, right in zip(month_raw, month))
        if score > best_score:
            best_score = score
            best_month = month
    return f"{day}{best_month}"


def looks_like_date(token: str) -> bool:
    return bool(SHORT_DATE_RE.match(normalize_date_token(token)))


def first_date_in_text(text: str) -> str:
    for token in normalize_ocr_text(text).split():
        normalized = normalize_date_token(token)
        if looks_like_date(normalized):
            return normalized
    return ""


def parse_amount(text: str) -> tuple[float | None, bool]:
    match = AMOUNT_RE.search(normalize_ocr_text(text).rstrip("."))
    if not match:
        return None, False

    amount_text = match.group(1)
    if "." in amount_text:
        normalized = amount_text.replace(",", "")
    elif "," in amount_text:
        normalized = amount_text.replace(",", ".")
    elif amount_text.isdigit() and len(amount_text) >= 3:
        normalized = f"{amount_text[:-2]}.{amount_text[-2:]}"
    else:
        normalized = amount_text

    try:
        value = float(normalized)
    except ValueError:
        return None, False
    return value, bool(match.group(2))


def page_to_ocr_image(page: pdfplumber.page.Page):
    cropped_page = page.crop(
        (
            page.width * 0.03,
            page.height * 0.05,
            page.width * 0.97,
            page.height * 0.95,
        )
    )
    return cropped_page.to_image(resolution=220).original.convert("L")


def extract_ocr_rows(image, box: tuple[int, int, int, int], min_conf: float = 20.0) -> list[dict]:
    cropped = image.crop(box)
    data = pytesseract.image_to_data(
        cropped,
        lang="eng",
        config="--psm 6",
        output_type=pytesseract.Output.DICT,
    )

    words: list[dict] = []
    for i, raw_text in enumerate(data["text"]):
        text = normalize_ocr_text(raw_text)
        conf_text = str(data["conf"][i]).strip()
        conf = float(conf_text) if conf_text and conf_text != "-1" else -1.0
        if not text or conf < min_conf:
            continue
        words.append(
            {
                "text": text,
                "x": box[0] + int(data["left"][i]),
                "y": box[1] + int(data["top"][i]),
            }
        )

    grouped: list[list[dict]] = []
    for word in sorted(words, key=lambda item: (item["y"], item["x"])):
        if not grouped or abs(word["y"] - grouped[-1][0]["y"]) > 10:
            grouped.append([word])
        else:
            grouped[-1].append(word)

    rows: list[dict] = []
    for items in grouped:
        items = sorted(items, key=lambda item: item["x"])
        rows.append(
            {
                "y": round(sum(item["y"] for item in items) / len(items)),
                "text": normalize_ocr_text(" ".join(item["text"] for item in items)),
            }
        )
    return rows


def nearest_row_text(rows: list[dict], y: int, tolerance: int = 12) -> str:
    match = next((row for row in rows if abs(row["y"] - y) <= tolerance), None)
    return match["text"] if match else ""


def build_hsbc_record(posting_date: str, transaction_date: str, details: str, amount_text: str) -> dict | None:
    amount, is_credit = parse_amount(amount_text)
    if amount is None:
        return None

    details_text = normalize_ocr_text(details).rstrip(".,")
    return {
        "Posting date": normalize_date_token(posting_date),
        "transection Date": normalize_date_token(transaction_date),
        "Transection details": details_text,
        "Amount (AED)": -abs(amount) if is_credit else amount,
    }


def extract_records_from_ocr_layout(page: pdfplumber.page.Page) -> list[dict]:
    image = page_to_ocr_image(page)
    width, height = image.size

    posting_rows = extract_ocr_rows(image, (0, 0, int(width * 0.14), height))
    transaction_rows = extract_ocr_rows(image, (int(width * 0.14), 0, int(width * 0.24), height))
    detail_rows = extract_ocr_rows(image, (int(width * 0.22), 0, int(width * 0.82), height))
    amount_rows = extract_ocr_rows(image, (int(width * 0.80), 0, width, height))

    summary_cutoffs = [row["y"] for row in detail_rows if any(marker in row["text"] for marker in SUMMARY_MARKERS)]
    bottom_cutoff = min(summary_cutoffs) if summary_cutoffs else height

    amount_rows = [row for row in amount_rows if 150 <= row["y"] < bottom_cutoff]
    detail_rows = [row for row in detail_rows if 150 <= row["y"] < bottom_cutoff]
    posting_rows = [row for row in posting_rows if row["y"] < bottom_cutoff]
    transaction_rows = [row for row in transaction_rows if row["y"] < bottom_cutoff]

    records: list[dict] = []
    current_record: dict | None = None
    matched_detail_ys: set[int] = set()

    for amount_row in amount_rows:
        amount_text = amount_row["text"]
        if "AMOUNT" in amount_text:
            continue

        detail_text = nearest_row_text(detail_rows, amount_row["y"])
        if not detail_text or "OPENING BALANCE" in detail_text:
            current_record = None
            continue

        posting_date = first_date_in_text(nearest_row_text(posting_rows, amount_row["y"]))
        transaction_date = first_date_in_text(nearest_row_text(transaction_rows, amount_row["y"]))

        if not posting_date and current_record is not None:
            posting_date = current_record["Posting date"]
        if not transaction_date and current_record is not None:
            transaction_date = current_record["transection Date"]
        if not posting_date:
            posting_date = transaction_date
        if not transaction_date:
            transaction_date = posting_date

        record = build_hsbc_record(posting_date, transaction_date, detail_text, amount_text)
        if not record:
            continue

        # Skip junk rows — detail is just symbols/numbers with no alphabetic content
        detail_alpha = re.sub(r"[^A-Z]", "", record["Transection details"])
        if len(detail_alpha) < 3:
            continue

        if not looks_like_date(record["Posting date"]) or not looks_like_date(record["transection Date"]):
            continue

        records.append(record)
        current_record = record
        matched_detail_ys.add(amount_row["y"])

    for detail_row in detail_rows:
        if detail_row["y"] in matched_detail_ys:
            continue
        if any(marker in detail_row["text"] for marker in SUMMARY_MARKERS):
            continue
        if current_record is None:
            continue
        if looks_like_date(detail_row["text"].split()[0] if detail_row["text"].split() else ""):
            continue
        if any(abs(detail_row["y"] - row["y"]) <= 12 for row in amount_rows):
            continue
        current_record["Transection details"] = normalize_ocr_text(
            f"{current_record['Transection details']} {detail_row['text']}"
        )

    cleaned_records = [
        record
        for record in records
        if record["Transection details"]
        and record["Amount (AED)"] != 0
        and "OPENING BALANCE" not in record["Transection details"]
        and "ACCOUNT SUMMARY" not in record["Transection details"]
        and "PAYMENT SUMMARY" not in record["Transection details"]
        and "SPECIFIED ACCOUNT WILL BE DEBITED" not in record["Transection details"]
        and "=" not in record["Transection details"]
    ]
    return cleaned_records


def ensure_hsbc_columns(df: pd.DataFrame) -> pd.DataFrame:
    for column in HSBC_COLUMN_ORDER:
        if column not in df.columns:
            df[column] = ""
    return df[HSBC_COLUMN_ORDER]


def run_hsbc_extraction(pdf_source: str | Path | BinaryIO) -> tuple[pd.DataFrame, dict]:
    all_records: list[dict] = []

    with pdfplumber.open(pdf_source) as pdf:
        for page in pdf.pages:
            all_records.extend(extract_records_from_ocr_layout(page))

    df = pd.DataFrame(all_records)
    if df.empty:
        return ensure_hsbc_columns(df), {}

    df = ensure_hsbc_columns(df)
    df = df.drop_duplicates().reset_index(drop=True)
    return df, {}


def write_hsbc_excel(df: pd.DataFrame, output_target: str | Path | BinaryIO) -> None:
    export_df = ensure_hsbc_columns(df.copy())

    with pd.ExcelWriter(output_target, engine="openpyxl") as writer:
        export_df.to_excel(writer, index=False, sheet_name="Transactions")
        ws = writer.sheets["Transactions"]

        for col, width in {"A": 16, "B": 18, "C": 56, "D": 18}.items():
            ws.column_dimensions[col].width = width

        header_fill = PatternFill("solid", fgColor="C8102E")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_fill = PatternFill("solid", fgColor="FFF8F8")
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 5):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill
            ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=4).number_format = '#,##0.00;#,##0.00 "CR"'


def hsbc_dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    write_hsbc_excel(df, buffer)
    buffer.seek(0)
    return buffer.getvalue()
