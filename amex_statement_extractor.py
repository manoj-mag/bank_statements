"""
AMEX Platinum Card (AED) statement extractor with optional header hints.

This module still works as a CLI script, and it now also exposes reusable
functions for a lightweight web interface.
"""

from __future__ import annotations

import re
import sys
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill

PDF_FILE = "STATEMENT.pdf"
OUTPUT_FILE = "amex_transactions.xlsx"

COLUMN_ORDER = [
    "Transaction Date",
    "Posting Date",
    "Details",
    "Non AED Spending",
    "Amount in AED",
]

DEFAULT_HEADER_ALIASES = {
    "Transaction Date": [
        "transaction date",
        "trans date",
        "date",
        "txn date",
    ],
    "Posting Date": [
        "posting date",
        "post date",
        "value date",
        "book date",
    ],
    "Details": [
        "details",
        "description",
        "merchant",
        "transaction details",
        "narration",
        "particulars",
    ],
    "Non AED Spending": [
        "non aed spending",
        "foreign amount",
        "foreign currency",
        "original amount",
        "billing amount foreign",
    ],
    "Amount in AED": [
        "amount in aed",
        "amount",
        "debit",
        "credit",
        "transaction amount",
        "billing amount",
        "local amount",
    ],
}


DATE_RE = re.compile(
    r"\b(\d{2}[/-][A-Za-z]{3}[/-]\d{4}|\d{2}[/-]\d{2}[/-]\d{4})\b"
)
AMOUNT_RE = re.compile(r"([\d,]+\.\d{2})\s*(CR)?")
FOREIGN_RE = re.compile(r"([\d,]+\.\d{2})\s+([A-Z]{3})\b")
REF_RE = re.compile(r"^Reference\s*:", re.IGNORECASE)
FX_RE = re.compile(r"^FX rate", re.IGNORECASE)

SKIP_PATTERNS = [
    re.compile(p, re.IGNORECASE)
    for p in [
        r"^Transaction Date",
        r"^Posting Date",
        r"^Details",
        r"^Non AED",
        r"^Amount in AED",
        r"^New Transactions For",
        r"^Card Account Number",
        r"^THANK YOU FOR USING",
        r"^IN RECONCILING",
        r"^Previous Balance",
        r"^New Credits",
        r"^New Debits",
        r"^New Balance",
        r"^Due Date",
        r"^Statement of Account",
        r"^The Platinum Card",
        r"^Prepared for",
        r"^Membership Number",
        r"^Statement date",
        r"^Page\s+\d",
        r"^American Express",
        r"^Customer Care",
        r"^Please pay",
        r"^Minimum Payment",
        r"^Closing Date",
        r"^Total due",
        r"^\*+",
    ]
]

SKIP_CONTAINS = [
    "transaction date posting date",
    "details non aed spending amount in aed",
    "new transactions for",
    "card account number",
    "membership number",
    "statement date",
]


def normalize_header_label(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(text).lower()).strip()


def merge_header_aliases(custom_aliases: dict[str, str] | None = None) -> dict[str, list[str]]:
    merged = {key: list(values) for key, values in DEFAULT_HEADER_ALIASES.items()}
    if not custom_aliases:
        return merged

    for key, extra_text in custom_aliases.items():
        if key not in merged or not extra_text:
            continue
        extras = [
            alias.strip().lower()
            for alias in re.split(r"[\n,;]+", extra_text)
            if alias.strip()
        ]
        merged[key].extend(extras)

    return merged


def is_skip_line(line: str) -> bool:
    line = line.strip()
    if not line:
        return True

    compact_line = re.sub(r"\s+", " ", line).strip().lower()
    if any(text in compact_line for text in SKIP_CONTAINS):
        return True

    return any(pat.search(line) for pat in SKIP_PATTERNS)


def parse_amount(text: str) -> tuple[float | None, bool]:
    text = str(text)
    matches = list(AMOUNT_RE.finditer(text))
    if not matches:
        return None, False
    match = matches[-1]
    value = float(match.group(1).replace(",", ""))
    lowered = text.lower()
    is_credit = bool(match.group(2)) or value < 0 or "(" in text or lowered.endswith(" cr")
    if "-" in text and match.start() > 0 and text[match.start() - 1] == "-":
        is_credit = True
    return value, is_credit


def extract_foreign(text: str) -> str:
    match = FOREIGN_RE.search(str(text))
    if match:
        return f"{match.group(1)} {match.group(2)}"
    return ""


def remove_last_amount(text: str) -> str:
    matches = list(AMOUNT_RE.finditer(str(text)))
    if not matches:
        return str(text).strip()
    last = matches[-1]
    return f"{text[:last.start()]}{text[last.end():]}".strip()


def strip_foreign_amount(text: str) -> str:
    return FOREIGN_RE.sub("", str(text)).strip()


def normalize_details_lines(lines: list[str]) -> str:
    cleaned = [line.strip() for line in lines if line and str(line).strip()]
    return "\n".join(cleaned)


def row_has_text(row: list[str]) -> bool:
    return any(str(cell).strip() for cell in row)


def detect_header_mapping(row: list[str], header_aliases: dict[str, list[str]]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized_cells = [normalize_header_label(cell) for cell in row]

    for idx, cell in enumerate(normalized_cells):
        if not cell:
            continue
        for canonical, aliases in header_aliases.items():
            if canonical in mapping:
                continue
            normalized_aliases = {normalize_header_label(alias) for alias in aliases}
            if cell in normalized_aliases or any(alias and alias in cell for alias in normalized_aliases):
                mapping[canonical] = idx
                break

    return mapping


def build_record_from_header_row(
    row: list[str],
    mapping: dict[str, int],
) -> dict | None:
    def get_value(column_name: str) -> str:
        idx = mapping.get(column_name)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx]).strip()

    txn_date = get_value("Transaction Date")
    if not DATE_RE.search(txn_date):
        return None

    post_date = get_value("Posting Date")
    details = normalize_details_lines(get_value("Details").splitlines())
    foreign = get_value("Non AED Spending")
    amount_text = get_value("Amount in AED")
    amount, is_credit = parse_amount(amount_text)

    if amount is None:
        amount = 0.0
        sign_hint = amount_text.strip().lower()
        if sign_hint.startswith("-"):
            amount = 0.0
            is_credit = True

    if not foreign:
        foreign = extract_foreign(details)
        if foreign:
            details = strip_foreign_amount(details)

    signed_amount = -abs(amount or 0.0) if is_credit else (amount or 0.0)
    return {
        "Transaction Date": DATE_RE.search(txn_date).group(1),
        "Posting Date": DATE_RE.search(post_date).group(1) if DATE_RE.search(post_date) else post_date,
        "Details": details,
        "Non AED Spending": foreign,
        "Amount in AED": signed_amount,
    }


def extract_transactions_from_table(
    table: list[list[str | None]],
    header_aliases: dict[str, list[str]],
) -> list[dict]:
    records: list[dict] = []
    header_mapping: dict[str, int] | None = None

    for raw_row in table:
        row = [str(cell).strip() if cell else "" for cell in raw_row]
        if not row_has_text(row):
            continue

        if header_mapping is None:
            possible_mapping = detect_header_mapping(row, header_aliases)
            if "Transaction Date" in possible_mapping and "Amount in AED" in possible_mapping:
                header_mapping = possible_mapping
                continue

        if header_mapping is not None:
            record = build_record_from_header_row(row, header_mapping)
            if record:
                records.append(record)
            continue

        if not DATE_RE.match(row[0]):
            continue

        txn_date = DATE_RE.match(row[0]).group(1)
        post_date = row[1] if len(row) > 1 else ""
        details = row[2] if len(row) > 2 else ""
        foreign = row[3] if len(row) > 3 else ""
        amount_text = row[4] if len(row) > 4 else ""
        amount, is_credit = parse_amount(amount_text)
        foreign_clean = extract_foreign(foreign) or foreign.strip()
        signed_amount = -(amount or 0.0) if is_credit else (amount or 0.0)

        records.append(
            {
                "Transaction Date": txn_date,
                "Posting Date": post_date,
                "Details": normalize_details_lines(details.splitlines()),
                "Non AED Spending": foreign_clean,
                "Amount in AED": signed_amount,
            }
        )

    return records


def extract_transactions_from_text(text: str) -> list[dict]:
    lines = text.splitlines()
    records = []
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        if is_skip_line(line):
            i += 1
            continue

        date_match = DATE_RE.match(line)
        if date_match:
            txn_date = date_match.group(1)
            remainder = line[date_match.end() :].strip()

            post_match = DATE_RE.match(remainder)
            post_date = ""
            if post_match:
                post_date = post_match.group(1)
                remainder = remainder[post_match.end() :].strip()

            details_lines = []
            foreign_amt = ""
            aed_amount = None
            is_credit = False

            if remainder:
                foreign_amt = extract_foreign(remainder) or foreign_amt
                amount, is_credit = parse_amount(remainder)
                if amount is not None:
                    aed_amount = amount
                    clean = remove_last_amount(remainder)
                    if foreign_amt:
                        clean = strip_foreign_amount(clean)
                    if clean:
                        details_lines.append(clean)
                else:
                    details_lines.append(remainder)

            i += 1
            while i < len(lines):
                next_line = lines[i].strip()
                if not next_line:
                    i += 1
                    continue
                if DATE_RE.match(next_line) and not REF_RE.match(next_line):
                    break
                if is_skip_line(next_line):
                    break
                if REF_RE.match(next_line) or FX_RE.match(next_line):
                    details_lines.append(next_line)
                    i += 1
                    continue

                foreign_amt = extract_foreign(next_line) or foreign_amt
                amount, amount_is_credit = parse_amount(next_line)
                if amount is not None and aed_amount is None:
                    aed_amount = amount
                    is_credit = amount_is_credit
                    clean = remove_last_amount(next_line)
                else:
                    clean = remove_last_amount(next_line)

                if foreign_amt:
                    clean = strip_foreign_amount(clean)
                if clean and not REF_RE.match(clean) and not FX_RE.match(clean):
                    details_lines.append(clean)

                i += 1

            records.append(
                {
                    "Transaction Date": txn_date,
                    "Posting Date": post_date,
                    "Details": normalize_details_lines(details_lines),
                    "Non AED Spending": foreign_amt,
                    "Amount in AED": -(aed_amount or 0.0) if is_credit else (aed_amount or 0.0),
                }
            )
            continue

        i += 1

    return records


def extract_summary_from_text(text: str) -> dict:
    summary = {}
    patterns = {
        "Previous Balance": r"Previous Balance[^\d]*([\d,]+\.\d{2})\s*(CR)?",
        "New Credits": r"New Credits[^\d]*([\d,]+\.\d{2})\s*(CR)?",
        "New Debits": r"New Debits[^\d]*([\d,]+\.\d{2})\s*(CR)?",
        "New Balance": r"New Balance[^\d]*([\d,]+\.\d{2})\s*(CR)?",
    }
    for key, pattern in patterns.items():
        match = re.search(pattern, text, re.IGNORECASE)
        if not match:
            continue
        value = float(match.group(1).replace(",", ""))
        summary[key] = f"{value:,.2f}{' CR' if match.group(2) else ''}"
    return summary


def ensure_dataframe_columns(df: pd.DataFrame, column_order: list[str] | None = None) -> pd.DataFrame:
    order = column_order or COLUMN_ORDER
    for column in order:
        if column not in df.columns:
            df[column] = ""
    return df[order]


def run_extraction(
    pdf_source: str | Path | BinaryIO,
    custom_aliases: dict[str, str] | None = None,
) -> tuple[pd.DataFrame, dict]:
    all_records = []
    summary = {}
    header_aliases = merge_header_aliases(custom_aliases)

    with pdfplumber.open(pdf_source) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            page_text = page.extract_text() or ""

            if page_num == 1 and page_text:
                summary = extract_summary_from_text(page_text)

            page_records = []
            if tables:
                for table in tables:
                    page_records.extend(extract_transactions_from_table(table, header_aliases))

            if page_records:
                all_records.extend(page_records)
            else:
                all_records.extend(extract_transactions_from_text(page_text))

    df = pd.DataFrame(all_records)
    if df.empty:
        return ensure_dataframe_columns(df), summary

    df = ensure_dataframe_columns(df)
    df = df.drop_duplicates().reset_index(drop=True)
    return df, summary


def write_excel(df: pd.DataFrame, summary: dict, output_target: str | Path | BinaryIO) -> None:
    # Use the columns as they are in the dataframe
    columns = list(df.columns)
    num_cols = len(columns)

    # Try to find the amount column for formatting and total
    amount_col_idx = num_cols # Default to last column
    amount_col_letter = chr(64 + num_cols)
    for i, col in enumerate(columns, start=1):
        if "amount" in col.lower():
            amount_col_idx = i
            amount_col_letter = chr(64 + i)
            break

    with pd.ExcelWriter(output_target, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Transactions")
        ws = writer.sheets["Transactions"]

        # Default widths
        widths = ["A", "B", "C", "D", "E", "F", "G"]
        default_widths = [18, 16, 55, 18, 18, 18, 18]
        for i in range(num_cols):
            ws.column_dimensions[widths[i]].width = default_widths[i]

        header_fill = PatternFill("solid", fgColor="006FCF")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_fill = PatternFill("solid", fgColor="F5F9FF")
        for row_idx in range(2, ws.max_row + 1):
            amount_cell = ws.cell(row=row_idx, column=amount_col_idx)
            amount_cell.number_format = '#,##0.00;#,##0.00 "CR"'

            for col_idx in range(1, num_cols + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

            # Assuming column 3 is always details/description for wrap text
            if num_cols >= 3:
                ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True, vertical="top")
            
            # Center align other columns except details and amount
            for col_idx in range(1, num_cols + 1):
                if col_idx != 3 and col_idx != amount_col_idx:
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(horizontal="center")
            
            amount_cell.alignment = Alignment(horizontal="right")

        last = ws.max_row
        ws.cell(row=last + 2, column=1, value="TOTALS").font = Font(bold=True, size=10)
        total_cell = ws.cell(row=last + 2, column=amount_col_idx, value=f"=SUM({amount_col_letter}2:{amount_col_letter}{last})")
        total_cell.number_format = "#,##0.00"
        total_cell.font = Font(bold=True)

        if summary:
            ws2 = writer.book.create_sheet("Summary")
            ws2.column_dimensions["A"].width = 22
            ws2.column_dimensions["B"].width = 18
            ws2["A1"] = "Account Summary"
            ws2["A1"].font = Font(bold=True, size=12, color="006FCF")
            for row_idx, (key, value) in enumerate(summary.items(), start=3):
                ws2.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
                ws2.cell(row=row_idx, column=2, value=value)


def dataframe_to_excel_bytes(df: pd.DataFrame, summary: dict) -> bytes:
    buffer = BytesIO()
    write_excel(df, summary, buffer)
    buffer.seek(0)
    return buffer.getvalue()


def save_excel(df: pd.DataFrame, summary: dict, output_path: str) -> None:
    write_excel(df, summary, output_path)


def resolve_pdf_path(argv: list[str]) -> str:
    if len(argv) > 1:
        return argv[1]

    default_path = Path(PDF_FILE)
    if default_path.exists():
        return str(default_path)

    pdf_files = sorted(Path.cwd().glob("*.pdf"))
    if len(pdf_files) == 1:
        return str(pdf_files[0])

    if len(pdf_files) > 1:
        print("\nMultiple PDF files found in this folder:")
        for pdf_file in pdf_files:
            print(f"  - {pdf_file.name}")
        print('Run: python3 amex_statement_extractor.py "your-file.pdf"')
        sys.exit(1)

    return str(default_path)


def main() -> None:
    pdf_path = resolve_pdf_path(sys.argv)

    if not Path(pdf_path).exists():
        print(f"\nFile not found: {pdf_path}")
        print("Put your statement PDF in this folder, or pass the file path explicitly.")
        print('Example: python3 amex_statement_extractor.py "My Statement.pdf"')
        sys.exit(1)

    print(f"\n{'=' * 55}")
    print("AMEX Statement Extractor")
    print(f"File : {pdf_path}")
    print(f"{'=' * 55}\n")

    df, summary = run_extraction(pdf_path)

    if df.empty:
        print("\nNo transactions extracted.")
        print("If this is a scanned PDF, OCR support would need to be added first.")
        sys.exit(0)

    print(f"Transactions extracted : {len(df)}")
    print(f"Credits                : {(df['Amount in AED'] < 0).sum()}")
    print(f"Debits                 : {(df['Amount in AED'] >= 0).sum()}")

    if summary:
        print("\nAccount Summary:")
        for key, value in summary.items():
            print(f"  {key:<20} {value}")

    preview_df = df.copy()
    preview_df["Amount in AED"] = preview_df["Amount in AED"].map(
        lambda amt: f"{abs(amt):,.2f} CR" if amt < 0 else f"{amt:,.2f}"
    )
    print("\nPreview:")
    print(preview_df[["Transaction Date", "Details", "Amount in AED"]].head(6).to_string(index=False))

    save_excel(df, summary, OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"{'=' * 55}\n")


if __name__ == "__main__":
    main()
